import argparse
import json
import os
import sys
from datetime import datetime
from typing import Dict, List, Optional, Set, Tuple

try:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.table import Table, TableStyleInfo
except ImportError:  # pragma: no cover - runtime environment dependency
    print("Missing dependency: openpyxl. Install it with: pip install openpyxl")
    sys.exit(1)


HEADERS = [
    "Mod Name",
    "Description",
    "Detail",
    "Category",
    "Links",
    "File Name",
    "Updated At",
]

# Column where the unique categories list will be placed (K by default).
CATEGORIES_COL_INDEX = 11
# Fixed height for data rows to avoid auto-expanding on large Detail cells.
DATA_ROW_HEIGHT = 15


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Generate a Mods Excel report from an ATLauncher instance.json file."
    )
    parser.add_argument("instance_json", help="Path to instance.json")
    return parser.parse_args()


def load_instance(path: str) -> Dict:
    with open(path, "r", encoding="utf-8") as handle:
        return json.load(handle)


def iso_to_datetime(value: Optional[str]) -> Optional[datetime]:
    if not value:
        return None
    try:
        normalized = value[:-1] + "+00:00" if value.endswith("Z") else value
        parsed = datetime.fromisoformat(normalized)
        if parsed.tzinfo is not None:
            parsed = parsed.astimezone().replace(tzinfo=None)
        return parsed
    except ValueError:
        return None


def modrinth_link(project: Dict) -> Optional[str]:
    slug = project.get("slug") or project.get("id")
    if not slug:
        return None
    return f"https://modrinth.com/mod/{slug}"


def mod_display_name(mod: Dict, project: Optional[Dict]) -> str:
    if project:
        return project.get("title") or mod.get("name") or mod.get("file") or ""
    return mod.get("name") or mod.get("file") or ""


def build_rows(mods: List[Dict], mods_dir: str) -> Tuple[List[Dict], Set[str]]:
    rows = []
    categories = set()

    for mod in mods:
        project = mod.get("modrinthProject")
        link_url = None
        if project:
            description = project.get("description") or ""
            detail = project.get("body") or ""
            category_list = project.get("categories") or []
            category = ";".join([c for c in category_list if c])
            for cat in category_list:
                if cat:
                    categories.add(cat)
            link_url = modrinth_link(project)
            link_text = "Modrinth" if link_url else ""
            updated_raw = project.get("updated")
        else:
            description = "No Modrinth source available (untrusted source)"
            detail = ""
            category = ""
            link_text = "No Modrinth"
            updated_raw = None

        updated_dt = iso_to_datetime(updated_raw)
        updated_value = updated_dt if updated_dt else (updated_raw or "")

        rows.append(
            {
                "Mod Name": mod_display_name(mod, project),
                "Description": description,
                "Detail": detail,
                "Category": category,
                "Links": link_text,
                "Links_url": link_url,
                "File Name": mod.get("file") or "",
                "Updated At": updated_value,
            }
        )

    return rows, categories


def write_excel(
    output_path: str, rows: List[Dict], categories: Set[str], sheet_name: str = "Mods"
) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name
    ws.freeze_panes = "A2"

    # Header row
    for col_idx, header in enumerate(HEADERS, start=1):
        ws.cell(row=1, column=col_idx, value=header)

    wrap_columns = {"Description", "Detail", "Category"}

    # Data rows
    for row_idx, row in enumerate(rows, start=2):
        for col_idx, header in enumerate(HEADERS, start=1):
            value = row.get(header, "")
            cell = ws.cell(row=row_idx, column=col_idx, value=value)

            if header == "Links" and row.get("Links_url"):
                cell.hyperlink = row["Links_url"]
                cell.style = "Hyperlink"

            if header == "Updated At" and isinstance(value, datetime):
                cell.number_format = "yyyy-mm-dd hh:mm:ss"

            if header in wrap_columns:
                cell.alignment = Alignment(wrap_text=True, vertical="top")
        ws.row_dimensions[row_idx].height = DATA_ROW_HEIGHT

    # Table (ListObject) for filters
    if rows:
        last_row = 1 + len(rows)
        table_ref = f"A1:{get_column_letter(len(HEADERS))}{last_row}"
        table = Table(displayName="ModsTable", ref=table_ref)
        style = TableStyleInfo(
            name="TableStyleMedium9",
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=False,
        )
        table.tableStyleInfo = style
        ws.add_table(table)

    # Reasonable column widths
    widths = [30, 50, 60, 30, 15, 40, 20]
    for idx, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = width

    # Unique categories list to the right of the table (column K by default)
    categories_col_letter = get_column_letter(CATEGORIES_COL_INDEX)
    ws.cell(row=1, column=CATEGORIES_COL_INDEX, value="Categorias")
    for offset, category in enumerate(sorted(categories, key=str.lower), start=2):
        ws.cell(row=offset, column=CATEGORIES_COL_INDEX, value=category)
    ws.column_dimensions[categories_col_letter].width = 25

    wb.save(output_path)


def main() -> int:
    args = parse_args()
    instance_path = os.path.abspath(args.instance_json)
    if not os.path.isfile(instance_path):
        print(f"File not found: {instance_path}")
        return 1

    instance = load_instance(instance_path)
    launcher = instance.get("launcher", {})
    mods = launcher.get("mods", [])

    instance_dir = os.path.dirname(instance_path)
    mods_dir = os.path.join(instance_dir, "mods")
    version = launcher.get("version") or "Unknown"
    output_path = os.path.join(instance_dir, f"Mods {version}.xlsx")

    rows, categories = build_rows(mods, mods_dir)
    write_excel(output_path, rows, categories)

    print(f"Excel report generated: {output_path}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
